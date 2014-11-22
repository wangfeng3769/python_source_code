#coding=utf-8
import xlrd
import re
import datetime
# wb = xlrd.open_workbook(u'E:/2013-3-14-满天星1.xls')
# for s in wb.sheets():
#     print s.name
# #    cinema_name = s.cell(0,4).value
# #    show_day = s.cell(0,16).value
#     for row in range(s.nrows):
#         try:
#             __s_date = datetime.date(1899, 12, 31).toordinal()-1
#             movie_name_mode = s.cell(row,5).value
#             hall_num = s.cell(row,1).value
#             show_date1 = s.cell(row,2).value
#             show_date = datetime.date.fromordinal(int(show_date1)+__s_date)
#             begin_time = s.cell(row,3).value
#             end_time = s.cell(row,4).value       # 结束时间
#             version = s.cell(row,6).value
#             price1 = s.cell(row,9).value
#             price = re.search(r'\d{1,3}',price1).group()
#             try:
#                 mode = re.search(u'[（\(][\s\S]*[\)）]',unicode(movie_name_mode)).group()
#                 movie_name =unicode(movie_name_mode).replace(mode,'')
#                 mode = mode[1:-1]
#             except Exception,e:
#                 mode = ''
#                 movie_name =movie_name_mode
#             print movie_name,mode
#             begin_time = xlrd.xldate_as_tuple(begin_time, 0)
#             end_time = xlrd.xldate_as_tuple(end_time, 0)
#             show_date1 = xlrd.xldate_as_tuple(show_date1,0)
#             begin_time1 = datetime.datetime(show_date1[0],show_date1[1],show_date1[2],begin_time[3],begin_time[4],begin_time[5])
#             end_time1 = datetime.datetime(show_date1[0],show_date1[1],show_date1[2],end_time[3],end_time[4],end_time[5])
#             if end_time1 < begin_time1:
#                 end_time1 += datetime.timedelta(days=1)
#             mins = (end_time1-begin_time1).seconds/60
#             begin_time1 = begin_time1.strftime('%H:%M')
#             end_time1 = end_time1.strftime('%H:%M')
#             print mins
#             content = dict(movie_name=movie_name,show_date=show_date,hall_num=hall_num,begin_time=begin_time1,end_time=end_time1,mins=mins,mode=mode,price=price)
#         except Exception,e:
#             print e

# wb = xlrd.open_workbook(u'E:/2013-3-14-满天星.xls')
from openpyxl import load_workbook
# wb = load_workbook(filename=u'E:/2013-3-14-满天星1.xls')
wb = load_workbook(filename=u'D:/苏州金逸亿象城IMAX影城-3-15-满天星1.xlsx')
for s in wb.sheets():
    print s.name
    #    cinema_name = s.cell(0,4).value
    #    show_day = s.cell(0,16).value
    for row in range(s.nrows):
        try:
            __s_date = datetime.date(1899, 12, 31).toordinal()-1
            movie_name_mode = s.cell(row,5).value
            hall_num = s.cell(row,1).value
            show_date1 = s.cell(row,2).value
            show_date = datetime.date.fromordinal(int(show_date1)+__s_date)
            begin_time = s.cell(row,3).value
            end_time = s.cell(row,4).value       # 结束时间
            version = s.cell(row,6).value
            price1 = s.cell(row,9).value
            price = re.search(r'\d{1,3}',price1).group()
            try:
                mode = re.search(u'[（\(][\s\S]*[\)）]',unicode(movie_name_mode)).group()
                movie_name =unicode(movie_name_mode).replace(mode,'')
                mode = mode[1:-1]
            except Exception,e:
                mode = ''
                movie_name =movie_name_mode
            print movie_name,mode
            begin_time = xlrd.xldate_as_tuple(begin_time, 0)
            end_time = xlrd.xldate_as_tuple(end_time, 0)
            show_date1 = xlrd.xldate_as_tuple(show_date1,0)
            begin_time1 = datetime.datetime(show_date1[0],show_date1[1],show_date1[2],begin_time[3],begin_time[4],begin_time[5])
            end_time1 = datetime.datetime(show_date1[0],show_date1[1],show_date1[2],end_time[3],end_time[4],end_time[5])
            if end_time1 < begin_time1:
                end_time1 += datetime.timedelta(days=1)
            mins = (end_time1-begin_time1).seconds/60
            begin_time1 = begin_time1.strftime('%H:%M')
            end_time1 = end_time1.strftime('%H:%M')
            print mins
            content = dict(movie_name=movie_name,show_date=show_date,hall_num=hall_num,begin_time=begin_time1,end_time=end_time1,mins=mins,mode=mode,price=price)
        except Exception,e:
            print e